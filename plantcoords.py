import numpy as np
import openpyxl, warnings, sys, ow
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=UserWarning)

openfile = str(sys.argv[1])

with open(openfile, 'r') as myfile:
	data = myfile.read()

data = data.split('["java.util.ArrayList/4159755760')
data = data[0][5:-1].split(",")
data = np.asarray(data).astype(float)

i=12
lat = []
lon = []

while i <= len(data):
	lat.append(data[i])
	lon.append(data[i+1])
	i+=15

wb = openpyxl.Workbook()
ws = wb.active
paste_range = ws["A1":"B"+str(len(lat))]


for row in np.arange(len(lat))+1: 
	ws.cell(column=1, row=row, value=lat[row-1])
	ws.cell(column=2, row=row, value=lon[row-1])

wb.save("coords2.xlsx")